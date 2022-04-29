# Busqueda_inf
from bs4 import BeautifulSoup as bs
from openpyxl import Workbook
from Direcciones_URL import Artista, path
import os
import requests
import re
import json

# Inicializando listas
Lista_correo=[]
Lista_naci=[]
Lista_tel=[]
Lista_head=[]
temp=[0,0,0]

# Verificar si existe la carpeta parrafos
try:
    os.mkdir(path+"/"+Artista+"/parrafos")
    print("Creando carpeta parrafos\n")
except:
    print("La carpeta parrafos ya existe")
    op=input("Desea sobreescribir los archivos? (y/n): ")
    op.lower()
    if op=="y" or op=="yes":
        print()
    else:
        exit()

# Verificar si existe la carpeta parrafosh
try:
    os.mkdir(path+"/"+Artista+"/parrafosh")
    print("Creando carpeta parrafosh\n")
except:
    print("La carpeta parrafosh ya existe")
    op=input("Desea sobreescribir los archivos? (y/n): ")
    op.lower()
    if op=="y" or op=="yes":
        print()
    else:
        exit()

# buscarP sirve para guardar el texto (todas las etiquetas (p) de un html en un documento de configuracion)
def buscarP(link, numLink, path):
    page=requests.get(link)
    soup=bs(page.content, "html.parser")
    po=soup.find_all("p")
    #print(len(po))
    #po es una lista con todos los parrafos
    fo=open(path+"/Texto_HTML_"+str(numLink)+".txt", "wb")
    for nParrafo in range(len(po)):
        potext=po[nParrafo].getText()
        fo.write((potext+"\n").encode("utf-8"))
        #print("potext")
    fo.close()

#buscarH sirve para guardar el texto (todas las etiquetas <h>) de un hatml en un documento
def buscarH(link, numLink, path):
    page=requests.get(link)
    soup=bs(page.content, "html.parser")
    po=soup.find_all(re.compile(r"^h[1-6]$"))
    #print(len(po))
    #po es una lista con todos los parrafos
    fo=open(path+"/Texto_HTML_H"+str(numLink)+".txt", "wb")
    for nParrafo in range(len(po)):
        potext=po[nParrafo].getText()
        fo.write((potext+"\n").encode("utf-8"))
        #print("potext")
    fo.close()

#Busca informacion con las expresiones regulares y guarda la informacion en un excel con el
def buscarInfdeP(path, num_Link):
    nom_doc=path+"/Texto_HTML_"+str(num_Link)+".txt"
    with open (nom_doc, "r", encoding="utf8") as file:
        for line in file:
            #print(line)
            br_regex=re.compile(r"\(\d(3)\) \d(3) \d(4)|\(\d(2)\) \d(4)-*\d(4)|\d(2)\s*\d(2)\s*\d(2)\s*\d(4)")
            tel=br_regex.findall(line)
            #print("tel", tel)
            if len(tel)!=0:
                for i in range(0, len(tel)):
                    Lista_tel.append(tel[i])

            br_correo=re.compile(r"correo: \w+@\w+.com|correo: \w+.\w+@\w+.com|\w+.\w+@\w+.com")
            correo=br_correo.findall(line)
            #print("correo", correo)
            if len(correo)!=0:
                for i in range(0, len(correo)):
                    Lista_correo.append(correo[i])

            br_naci=re.compile(r"\d+\s*\w+\s*\w+\s*\w+\s*\d+\s*nacimiento|Fecha de nacimiento: \d(2)/\d(2)/\d(4)|Fecha de nacimiento: \d(2) \w \d(4)|Birthday \d(2)/\d(2)?/\d(4)")
            nacimiento=br_naci.findall(line)
            #print("nacimiento", nacimiento)
            if (len(nacimiento))!=0:
                for i in range(0, len(nacimiento)):
                    Lista_naci.append(nacimiento[i])

#Busca informacion con las expresiones regulares y guarda la informacion en un excel con el nombre
def buscarInfdeH(path, num_Link):
    nom_doc=(path+"/Texto_HTML_H"+str(num_Link)+".txt")
    with open(nom_doc, "r", encoding="utf8") as file:
        for line in file:
            #print(line)
            br_regex=re.compile(r"Cristiano Ronaldo \w+ \w+ \w+ \w+ \w+ \w+ \w+ \w+ \w+ \w+| Cristiano Ronaldo \w+ \w+ \w+ \w+ \w+ \w+ \w+ \w+|Cristiano Ronaldo \w+ \w+ \w+ \w+ \w+ \w+ \w+|Cristiano Ronaldo \w+ \w+ \w+ \w+ \w+ \w+")
            head=br_regex.findall(line)
            #print("tel", tel)
            if len(head)!=0:
                for i in range(0, len(head)):
                    Lista_head.append(head[i])

#Busca la tempetura de 3 dias gracias a la api de OpenWeather
#def OpenWeather():
#    print("Buscando temperatura...\n")
#    url="https://api.openweathermap.org/data/2.5/onecall?lat=42.8333&lon12.8333&exclude=c"
#    response=requests.get(url)
#    datos=json.loads(response.content)
#    #print(datos)
#
#    temp[0]=datos["daily"] [0] ["temp"] ["day"]
#    #print("partido", temp[0])
#    temp[1]=datos["daily"] [1] ["temp"] ["day"]
#    #print("partido 2", temp[1])
#    temp[2]=datos["daily"] [2] ["temp"] ["day"]
    #print("partido 3", temp[2])

#Creacion de Excel y guardado de informacion de funciones anteriores
def Excel():
    #Se crea el Excel
    wb=Workbook()
    ws=wb.active
    ws.title="Datos"
#    ws1=wb.create_sheet("Clima")
    #Se marcan las celdas a utilizar
    celda1=ws["A1"]
    celda2=ws["B1"]
    celda3=ws["C1"]
    celda4=ws["D1"]
#    celda5=ws1["A1"]
#    celda6=ws1["A1"]
    #Le damos un valor a cada celda
    celda1.value="Telefonos"
    celda2.value="Correos electronicos"
    celda3.value="Fecha de nacimiento"
    celda4.value="Encabezados"
#    celda5.value="Dias"
#    celda6.value="Temperatura"
#    Lista_Dias=["6 de noviembre", "9 de noviembre", "10 de noviembre"]
#    Lista_Temp=[temp[0], temp[1], temp[2]]
    #print(Lista_Temp)

    #Lista de las listas XD
    Listas_Listas=[Lista_tel, Lista_correo, Lista_naci, Lista_head]
#    Listas_Listas2=[Lista_Dias, Lista_Temp]

    columna=1
    # Guardar la informacion de las listas en celdas
    for x in Listas_Listas:
        fila=2
        for y in x:
            d=ws.cell(row=fila, column=columna, value=y)
            fila+=1
        columna+=1
    columna=1


#    for x in Listas_Listas2:
#        fila=2
#        for y in x:
#            d=ws.cell(row=fila, column=columna, value=y)
#            fila+=1
#        columna+=1

    # Se verifica si existe ya el documento creado
    try:
        with open(path+"/"+Artista+"/"+Artista+".xlsx", "r") as f:
            if True:
                print("Ya existe el documento excel: "+Artista+".xlsx")
                op=input("Desea sobreescribir el archivo? (y/n): ")
                op.lower()
                if op=="y" or op=="yes":
                    print()
                else:
                    exit()
    except:
        print("Guardando archivo excel...\n")
        wb.save(path+"/"+Artista+"/"+Artista+".xlsx")
